"""
한국 주식시장 종목 스크리너 (KOSPI/KOSDAQ)
데이터 출처: 네이버금융 (시가총액 페이지 + 개별종목 시세) - 로그인/API키 불필요, 무료

스크리닝 조건
  --market-cap-min     시가총액 하한 (억원)
  --volume-min         기준일(당일) 거래량 하한 (주)
  --cum-decline-min    누적하락률 하한 (%, 예: 20 -> 최근 N거래일간 20% 이상 하락한 종목)
  --up-days-min        기간 내 상승일수 하한 (일)
  --consecutive-up-min 최근 연속 상승일수 하한 (일)
  --theme-keyword       이슈테마주 키워드 (네이버금융 테마 페이지 기준, 콤마로 여러개)
  --min-data-days      분석하한: 최소 거래 이력 일수 (신규상장 등 데이터 부족 종목 제외)

동작 방식
  1) 네이버금융 시가총액 페이지(전 종목)를 스캔해 시총/거래량 조건으로 1차 필터링
  2) 통과한 후보 종목만 개별 시세 이력을 조회해 누적등락률/상승일수/연속상승일수 계산
  3) (선택) 네이버금융 테마 페이지와 대조해 이슈테마주 태깅/필터링

사용 예시
  python stock_screener.py --market-cap-min 1000 --volume-min 500000 --up-days-min 4
  python stock_screener.py --cum-decline-min 20 --period 20
  python stock_screener.py --theme-keyword 2차전지,로봇
  python stock_screener.py --code 006400          (삼성SDI 개별 조회)
  python stock_screener.py --name 삼성SDI --period 20

설치
  pip install -r requirements.txt
"""

import argparse
import os
import sys
import time
import unicodedata
from datetime import datetime, timedelta

import pandas as pd
import requests
from bs4 import BeautifulSoup

HEADERS = {"User-Agent": "Mozilla/5.0"}


def _display_width(s: str) -> int:
    """터미널에서 한글은 2칸, 영문/숫자는 1칸을 차지하므로 실제 표시 폭을 계산한다."""
    return sum(2 if unicodedata.east_asian_width(ch) in ("W", "F") else 1 for ch in s)


def _pad(s: str, width: int) -> str:
    return " " * max(0, width - _display_width(s)) + s


def format_table(df: pd.DataFrame) -> str:
    """한글/영문 혼용 표를 터미널에서 오와 열이 맞게 정렬해 문자열로 만든다.
    (pandas to_string()은 글자 수 기준이라 한글이 섞이면 폭 2칸을 반영 못해 틀어진다.)"""
    if df.empty:
        return "(조건에 맞는 종목이 없습니다)"

    cols = list(df.columns)
    cell = {c: [("" if pd.isna(v) else str(v)) for v in df[c]] for c in cols}
    widths = {c: max([_display_width(c)] + [_display_width(v) for v in cell[c]]) for c in cols}

    lines = ["  ".join(_pad(c, widths[c]) for c in cols)]
    for i in range(len(df)):
        lines.append("  ".join(_pad(cell[c][i], widths[c]) for c in cols))
    return "\n".join(lines)


COLOR_COLS = {"등락률(%)", "기간누적등락률(%)", "평가손익", "손익률(%)"}
INT_COLS = {"시가총액(억)", "현재가", "종가", "거래량", "평균단가", "보유잔고", "매도가능",
            "매입금", "평가금", "평가손익", "기간거래량합계"}
PERCENT_COLS = {"등락률(%)", "기간누적등락률(%)", "손익률(%)"}


def _style_and_chart_sheet(ws, df):
    """워크시트 하나에 헤더 서식/색상/차트를 적용한다 (save_excel, save_excel_sheets 공용).
    호출 시점에 openpyxl이 이미 임포트되어 있다고 가정한다."""
    from openpyxl.chart import PieChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.marker import DataPoint
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.chart.text import RichText, Text
    from openpyxl.chart.title import Title
    from openpyxl.drawing.line import LineProperties
    from openpyxl.drawing.text import (CharacterProperties, Paragraph, ParagraphProperties,
                                        RegularTextRun, RichTextProperties)
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="404040")
    for col_idx in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for col_idx, col_name in enumerate(df.columns, start=1):
        letter = get_column_letter(col_idx)
        values = df[col_name]
        max_w = max([_display_width(str(col_name))] + [_display_width(str(v)) for v in values]) if len(values) else _display_width(str(col_name))
        ws.column_dimensions[letter].width = max_w + 2

        if col_name in PERCENT_COLS:
            fmt = "#,##0.00"
        elif col_name in INT_COLS:
            fmt = "#,##0"
        else:
            fmt = None

        for row_idx in range(2, len(values) + 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            if fmt:
                cell.number_format = fmt
            if col_name in COLOR_COLS and isinstance(cell.value, (int, float)):
                if cell.value > 0:
                    cell.font = Font(color="D60000")
                elif cell.value < 0:
                    cell.font = Font(color="0051C7")

    ws.freeze_panes = "A2"

    n_rows = len(df)
    cols = list(df.columns)
    if n_rows > 0 and "종목명" in cols and {"매입금", "평가금"} <= set(cols):
        name_idx = cols.index("종목명") + 1
        eval_idx = cols.index("평가금") + 1
        cats = Reference(ws, min_col=name_idx, min_row=2, max_row=n_rows + 1)

        def rich_title(text, size=1500, color="404040"):
            cp = CharacterProperties(sz=size, b=True, solidFill=color, latin=None, ea=None)
            run = RegularTextRun(t=text, rPr=cp)
            para = Paragraph(pPr=ParagraphProperties(defRPr=cp), r=[run])
            return Title(tx=Text(rich=RichText(p=[para])))

        def horizontal_labels(chart, color="404040", extra=None):
            """데이터 라벨 글자를 세로로 눕히지 않고 가로 방향으로 고정해서 표시한다."""
            dl = DataLabelList()
            dl.showVal = True
            dl.showCatName = True
            dl.showPercent = True
            dl.showLegendKey = False
            dl.showSerName = False
            dl.numFmt = "#,##0"
            dl.dLblPos = "outEnd"
            cp = CharacterProperties(sz=900, b=True, solidFill=color)
            body = RichTextProperties(rot=0, vert="horz")  # 글자 방향: 가로
            dl.txPr = RichText(bodyPr=body, p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
            if extra:
                extra(dl)
            chart.dataLabels = dl

        # 차트 1: 평가금 비중 (종목별 현재 평가금액 비율을 원형 그래프로)
        chart1 = PieChart()
        chart1.varyColors = True
        chart1.title = rich_title("평가금 비중 (종목별)")
        chart1.style = 2
        chart1.height, chart1.width = 11, 15

        data1 = Reference(ws, min_col=eval_idx, min_row=1, max_row=n_rows + 1)
        chart1.add_data(data1, titles_from_data=True)
        chart1.set_categories(cats)
        horizontal_labels(chart1)
        chart1.legend.position = "b"
        chart1.legend.overlay = False
        chart1.legend.txPr = RichText(
            bodyPr=RichTextProperties(rot=0, vert="horz"),
            p=[Paragraph(pPr=ParagraphProperties(defRPr=CharacterProperties(sz=900)), endParaRPr=CharacterProperties(sz=900))])

        anchor_col = get_column_letter(len(cols) + 2)
        ws.add_chart(chart1, f"{anchor_col}2")

        pnl_col = "평가손익" if "평가손익" in cols else None
        if pnl_col:
            pnl_idx = cols.index(pnl_col) + 1
            chart2 = PieChart()
            chart2.varyColors = True
            chart2.title = rich_title(f"{pnl_col} 비중  (▲ 이익 = 빨강  ·  ▼ 손실 = 파랑)")
            chart2.style = 2
            chart2.height, chart2.width = 11, 15
            chart2.legend = None

            data2 = Reference(ws, min_col=pnl_idx, min_row=1, max_row=n_rows + 1)
            chart2.add_data(data2, titles_from_data=True)
            chart2.set_categories(cats)

            def no_percent(dl):
                dl.showPercent = False  # 손익은 부호가 섞여 비율(%)이 왜곡되므로 금액만 표시
            horizontal_labels(chart2, extra=no_percent)

            pts = []
            for i in range(n_rows):
                val = df[pnl_col].iloc[i]
                if pd.notna(val) and val > 0:
                    color = "E5484D"     # 이익 = 세련된 레드
                elif pd.notna(val) and val < 0:
                    color = "3E7BFA"     # 손실 = 세련된 블루
                else:
                    color = "BFBFBF"
                gp = GraphicalProperties(solidFill=color)
                gp.line = LineProperties(solidFill="FFFFFF", w=19050)
                pts.append(DataPoint(idx=i, spPr=gp))
            chart2.series[0].data_points = pts

            gap_rows = 24 + max(0, n_rows - 5) * 1
            ws.add_chart(chart2, f"{anchor_col}{gap_rows}")


def _excel_available():
    try:
        import openpyxl  # noqa: F401
        return True
    except ImportError:
        return False


def save_excel_sheets(sheets, path) -> str:
    """(시트이름, DataFrame) 목록을 한 엑셀 파일에 시트별로 나눠 저장한다.
    관심종목/종목검색/종목조회/테마주처럼 여러 조회 결과를 하나의 파일로
    묶어서 검색이 끝나면 한 번에 보여줄 때 사용한다.
    파일이 이미 있으면 덮어쓰지 않고 새 시트로 이어 붙인다 (이력 누적).
    openpyxl이 없으면 시트마다 별도 CSV로 대신 저장한다."""
    sheets = [(name, df) for name, df in sheets if df is not None and not df.empty]
    if not sheets:
        raise ValueError("결과가 있는 항목이 하나도 없습니다.")

    if not _excel_available():
        print("openpyxl이 설치되어 있지 않아 CSV로 대신 저장합니다. (pip install openpyxl)", file=sys.stderr)
        base = os.path.splitext(path)[0]
        for name, df in sheets:
            df.to_csv(f"{base}_{name}.csv", index=False, encoding="utf-8-sig")
        return base

    file_exists = os.path.exists(path)
    writer_kwargs = {"engine": "openpyxl"}
    if file_exists:
        writer_kwargs.update(mode="a", if_sheet_exists="new")

    with pd.ExcelWriter(path, **writer_kwargs) as writer:
        for name, df in sheets:
            df.to_excel(writer, index=False, sheet_name=name)
            _style_and_chart_sheet(writer.sheets[name], df)

    return path


def save_excel(df: pd.DataFrame, path: str) -> str:
    """결과를 엑셀(.xlsx)로 저장한다 (단일 시트). 파일이 이미 있으면 덮어쓰지 않고
    실행 시각으로 이름 붙인 새 시트를 추가해서 이전 조회 이력이 한 파일 안에
    계속 누적되도록 한다 (같은 파일, 시트만 늘어남).
    헤더 굵게+색배경, 열 너비 자동 조정, 상승/이익은 빨강 하락/손실은 파랑으로 표시
    (국내 증권 앱 관행: 상승=빨강, 하락=파랑). openpyxl이 없으면 CSV로 대신 저장."""
    sheet_name = datetime.today().strftime("%Y-%m-%d_%H%M%S")
    return save_excel_sheets([(sheet_name, df)], path)


try:
    from pykrx import stock  # 개별 종목 시세 이력(네이버 기반)만 사용. 전종목 일괄조회는 자체 구현.
except ImportError:
    print("pykrx가 설치되어 있지 않습니다. 다음 명령으로 설치하세요:\n  pip install pykrx", file=sys.stderr)
    sys.exit(1)


def _to_float(s: str):
    s = s.replace(",", "").replace("%", "").strip()
    if s in ("", "-"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def fetch_market_summary(market: str) -> pd.DataFrame:
    """네이버금융 시가총액 페이지에서 전종목 현재가/등락률/시가총액/거래량을 수집한다.
    market: 'KOSPI'(sosok=0) 또는 'KOSDAQ'(sosok=1)
    """
    sosok = 0 if market == "KOSPI" else 1
    base_url = f"https://finance.naver.com/sise/sise_market_sum.naver?sosok={sosok}&page=1"
    r = requests.get(base_url, headers=HEADERS, timeout=10)
    r.encoding = "euc-kr"
    soup = BeautifulSoup(r.text, "html.parser")
    last_link = soup.select_one("td.pgRR a")
    last_page = 1
    if last_link and "page=" in last_link["href"]:
        last_page = int(last_link["href"].split("page=")[-1])

    rows = []
    for page in range(1, last_page + 1):
        url = f"https://finance.naver.com/sise/sise_market_sum.naver?sosok={sosok}&page={page}"
        resp = requests.get(url, headers=HEADERS, timeout=10)
        resp.encoding = "euc-kr"
        s = BeautifulSoup(resp.text, "html.parser")
        table = s.select_one("table.type_2")
        if table is None:
            continue
        for tr in table.select("tr"):
            a = tr.select_one("a.tltle")
            if not a:
                continue
            code = a["href"].split("code=")[-1]
            tds = [td.get_text(strip=True) for td in tr.select("td")]
            # [순위, 종목명, 현재가, 전일비, 등락률, 액면가, 시가총액(억), 상장주식수(천주), 외국인비율, 거래량, PER, ROE]
            if len(tds) < 12:
                continue
            rows.append({
                "종목코드": code,
                "종목명": a.get_text(strip=True),
                "현재가": _to_float(tds[2]),
                "등락률(%)": _to_float(tds[4]),
                "시가총액(억)": _to_float(tds[6]),
                "거래량": _to_float(tds[9]),
            })
        time.sleep(0.1)

    df = pd.DataFrame(rows).drop_duplicates(subset="종목코드").reset_index(drop=True)
    return df


def trading_days_from(n: int) -> str:
    """오늘 기준 약 n 거래일 이전 날짜(YYYYMMDD, 대략치). 캘린더로 넉넉히 잡음."""
    return (datetime.today() - timedelta(days=int(n * 1.6) + 5)).strftime("%Y%m%d")


def compute_history_stats(tickers, fromdate: str, todate: str, pause: float = 0.15):
    """후보 종목에 대해서만 개별 시세 이력을 조회해
    (기간누적등락률, 기간상승일수, 연속상승일수, 데이터일수)를 계산한다.
    네이버 개별종목 시세를 사용하므로 로그인 불필요."""
    result = {}
    total = len(tickers)
    for i, t in enumerate(tickers, 1):
        try:
            daily = stock.get_market_ohlcv_by_date(fromdate, todate, t)
        except Exception:
            result[t] = None
            continue

        if daily is None or daily.empty or "종가" not in daily.columns:
            result[t] = None
            continue

        closes = daily["종가"]
        cum_return = (closes.iloc[-1] / closes.iloc[0] - 1) * 100 if closes.iloc[0] else None

        if "등락률" in daily.columns:
            rets = daily["등락률"]
        else:
            rets = closes.pct_change().fillna(0) * 100

        up_days = int((rets > 0).sum())
        streak = 0
        for r in rets.iloc[::-1]:
            if r > 0:
                streak += 1
            else:
                break

        result[t] = {
            "종가": int(closes.iloc[-1]),
            "기간누적등락률(%)": round(cum_return, 2) if cum_return is not None else None,
            "기간상승일수": up_days,
            "연속상승일수": streak,
            "데이터일수": len(daily),
        }

        if pause:
            time.sleep(pause)
        if total > 30 and i % 30 == 0:
            print(f"  ... 진행 {i}/{total}", file=sys.stderr)

    return result


def read_watchlist(path: str):
    """관심종목 텍스트 파일을 읽어 (종목코드 목록, 종목명 목록)을 반환한다.
    한 줄에 하나씩. '#'으로 시작하는 줄과 빈 줄은 무시.
    6자리 숫자면 종목코드로, 그 외는 종목명으로 취급."""
    codes, names = [], []
    with open(path, "r", encoding="utf-8-sig") as f:
        for line in f:
            line = line.split("#", 1)[0].strip()
            if not line:
                continue
            if line.isdigit() and len(line) == 6:
                codes.append(line)
            else:
                names.append(line)
    return codes, names


_theme_map_cache = None


def fetch_theme_map():
    """네이버금융 테마 페이지(전 페이지)에서 테마명 -> 소속 종목명 매핑을 수집한다.
    비공식 페이지 구조를 파싱하므로, 네이버 페이지 개편 시 동작하지 않을 수 있음.
    테마 수가 많아(약 250개+) 수 분이 걸릴 수 있다.
    같은 프로세스 안에서는 결과를 메모리에 캐싱해서, 통합조회처럼 테마를
    여러 번 조회할 때 매번 다시 긁어오지 않고 한 번만 가져온다."""
    global _theme_map_cache
    if _theme_map_cache is not None:
        return _theme_map_cache

    first = requests.get("https://finance.naver.com/sise/theme.naver", headers=HEADERS, timeout=10)
    first.encoding = "euc-kr"
    soup = BeautifulSoup(first.text, "html.parser")
    last_link = soup.select_one("table.Nnavi td.pgRR a")
    last_page = 1
    if last_link and "page=" in last_link["href"]:
        last_page = int(last_link["href"].split("page=")[-1])

    theme_links = []
    for page in range(1, last_page + 1):
        url = f"https://finance.naver.com/sise/theme.naver?page={page}"
        resp = requests.get(url, headers=HEADERS, timeout=10)
        resp.encoding = "euc-kr"
        s = BeautifulSoup(resp.text, "html.parser")
        for a in s.select("table.type_1 a"):
            href = a.get("href", "")
            if "sise/sise_group_detail" in href:
                theme_links.append((a.text.strip(), "https://finance.naver.com" + href))
        time.sleep(0.1)

    print(f"  테마 {len(theme_links)}개 수집, 종목 매핑 중...", file=sys.stderr)
    stock_to_themes = {}
    for i, (theme_name, url) in enumerate(theme_links, 1):
        try:
            r = requests.get(url, headers=HEADERS, timeout=10)
            r.encoding = "euc-kr"
            s = BeautifulSoup(r.text, "html.parser")
            for a in s.select("table.type_5 a"):
                href = a.get("href", "")
                if "code=" in href:
                    name = a.text.strip()
                    if name:
                        stock_to_themes.setdefault(name, set()).add(theme_name)
            time.sleep(0.15)
        except Exception:
            continue

        if i % 50 == 0:
            print(f"  ... 테마 매핑 {i}/{len(theme_links)}", file=sys.stderr)

    _theme_map_cache = stock_to_themes
    return stock_to_themes


def run_screen(args):
    """전체 스크리닝 로직. args는 argparse.Namespace 이거나 아래와 동일한 속성을
    가진 객체(예: types.SimpleNamespace)면 된다 - 웹앱 등 다른 인터페이스에서도
    이 함수 하나로 CLI와 완전히 동일한 스크리닝을 재사용할 수 있다.

    필요 속성: period, market, market_cap_min, volume_min, cum_decline_min,
    up_days_min, consecutive_up_min, theme_keyword, min_data_days, code, name,
    search, watchlist, buy_price, quantity

    반환: 결과 pandas.DataFrame (표시용 컬럼만 정리된 상태)"""
    min_data_days = args.min_data_days or args.period
    from_date = trading_days_from(args.period)
    to_date = datetime.today().strftime("%Y%m%d")

    markets = ["KOSPI", "KOSDAQ"] if args.market == "ALL" else [args.market]

    print(f"[조회 기간] {from_date} ~ {to_date}  [시장] {markets}", file=sys.stderr)
    print("네이버금융 시가총액 페이지에서 전종목 스캔 중... (KOSPI 약 50p, KOSDAQ 약 55p)", file=sys.stderr)
    frames = []
    for m in markets:
        frames.append(fetch_market_summary(m))
    df = pd.concat(frames, ignore_index=True)
    print(f"전체 유니버스: {len(df)}종목", file=sys.stderr)

    single_lookup = bool(args.code or args.name or args.watchlist or args.search)
    if single_lookup:
        codes = [c.strip() for c in args.code.split(",")] if args.code else []
        names = [n.strip() for n in args.name.split(",")] if args.name else []
        if args.watchlist:
            try:
                wl_codes, wl_names = read_watchlist(args.watchlist)
            except FileNotFoundError:
                print(f"관심종목 파일을 찾을 수 없습니다: {args.watchlist}", file=sys.stderr)
                raise

            codes += wl_codes
            names += wl_names

        if args.watchlist and not codes and not names and not args.search:
            raise ValueError(f"{args.watchlist} 파일이 비어있습니다. 종목을 추가하거나 직접 입력해 주세요.")

        mask = df["종목코드"].isin(codes) | df["종목명"].isin(names)
        if args.search:
            keywords = [k.strip() for k in args.search.split(",") if k.strip()]
            for kw in keywords:
                mask = mask | df["종목명"].str.contains(kw, case=False, na=False) \
                             | df["종목코드"].str.contains(kw, case=False, na=False)
        df = df[mask]

        if df.empty:
            raise ValueError("해당 종목코드/종목명을 찾을 수 없습니다. 코드나 이름을 확인하세요.")
        print(f"지정 종목 조회: {len(df)}종목 ({', '.join(df['종목명'])})", file=sys.stderr)
    else:
        if args.market_cap_min:
            df = df[df["시가총액(억)"].fillna(0) >= args.market_cap_min]
        if args.volume_min:
            df = df[df["거래량"].fillna(0) >= args.volume_min]
        print(f"1차 필터(시총/거래량) 통과: {len(df)}종목", file=sys.stderr)

    need_history = single_lookup or (
        args.cum_decline_min is not None
        or args.up_days_min is not None
        or args.consecutive_up_min is not None
    )
    if need_history:
        candidates = df["종목코드"].tolist()
        print(f"기간 시세 이력 계산 대상: {len(candidates)}종목 (종목당 조회, 시간이 걸릴 수 있습니다)", file=sys.stderr)
        stats = compute_history_stats(candidates, from_date, to_date)

        df["종가"] = df["종목코드"].map(lambda t: (stats.get(t) or {}).get("종가"))
        df["기간누적등락률(%)"] = df["종목코드"].map(lambda t: (stats.get(t) or {}).get("기간누적등락률(%)"))
        df["기간상승일수"] = df["종목코드"].map(lambda t: (stats.get(t) or {}).get("기간상승일수"))
        df["연속상승일수"] = df["종목코드"].map(lambda t: (stats.get(t) or {}).get("연속상승일수"))
        df["데이터일수"] = df["종목코드"].map(lambda t: (stats.get(t) or {}).get("데이터일수"))

        df = df[df["데이터일수"].fillna(0) >= min_data_days]
        if args.cum_decline_min is not None:
            df = df[df["기간누적등락률(%)"].fillna(0) <= -abs(args.cum_decline_min)]
        if args.up_days_min is not None:
            df = df[df["기간상승일수"].fillna(-1) >= args.up_days_min]
        if args.consecutive_up_min is not None:
            df = df[df["연속상승일수"].fillna(-1) >= args.consecutive_up_min]

    if args.theme_keyword:
        keywords = [k.strip() for k in args.theme_keyword.split(",") if k.strip()]
        print(f"이슈테마주 매칭 중 (네이버금융 테마 기준, 키워드: {keywords})...", file=sys.stderr)
        try:
            theme_map = fetch_theme_map()
            df["테마"] = df["종목명"].map(lambda n: ", ".join(sorted(theme_map.get(n, []))))
            mask = df["테마"].apply(lambda themes: any(k in themes for k in keywords))
            df = df[mask]
        except Exception as e:
            print(f"테마 조회 실패 (네이버 페이지 구조 변경 가능성): {e}", file=sys.stderr)

    if "종가" not in df.columns:
        df["종가"] = df["현재가"]  # 개별 이력 조회를 안 한 경우 현재가로 대체(당일 종가와 사실상 동일)

    if args.buy_price:
        qty = args.quantity or 1
        df["평균단가"] = args.buy_price
        df["보유잔고"] = qty
        df["매도가능"] = qty  # 대주/담보 등 잠긴 수량을 별도로 추적하지 않으므로 보유잔고와 동일
        df["매입금"] = round(args.buy_price * qty)
        df["평가금"] = (df["현재가"] * qty).round(0)
        df["평가손익"] = (df["평가금"] - df["매입금"]).round(0)
        df["손익률(%)"] = ((df["현재가"] - args.buy_price) / args.buy_price * 100).round(2)

    df = df.sort_values("시가총액(억)", ascending=False)
    df["조회일자"] = datetime.today().strftime("%Y-%m-%d")

    if args.buy_price:
        keep_cols = [c for c in [
            "조회일자", "종목코드", "종목명", "평가손익", "손익률(%)", "매입금", "평가금",
            "매도가능", "평균단가", "보유잔고", "현재가",
        ] if c in df.columns]
    else:
        keep_cols = [c for c in [
            "조회일자", "종목코드", "종목명", "현재가", "종가", "등락률(%)", "거래량", "테마",
        ] if c in df.columns]
    return df[keep_cols].reset_index(drop=True)


DEFAULT_SCREEN_ARGS = dict(
    period=20, market="ALL", market_cap_min=0, volume_min=0,
    cum_decline_min=None, up_days_min=None, consecutive_up_min=None,
    theme_keyword=None, min_data_days=None, code=None, name=None,
    search=None, watchlist=None, buy_price=None, quantity=None,
)


def run_combined(wl_input=None, watchlist_path="관심종목.txt", search_kw=None,
                  lookup=None, theme_kw=None, buy_price=None, quantity=None, period=20):
    """관심종목/종목검색/종목조회/테마주 네 가지를 순서대로 조회해서
    (시트이름, DataFrame) 목록으로 반환한다. 값을 안 준 항목은 건너뛴다.
    하나라도 실패해도(종목 없음 등) 나머지는 계속 진행한다.

    테마주는 이 기능의 중심이라 콤마로 여러 테마를 넣으면(예: "2차전지,로봇")
    하나로 합쳐서 찾지 않고 테마별로 각각 시트를 따로 만든다
    (테마_2차전지, 테마_로봇, ...) - 테마마다 결과를 비교해서 보기 좋게."""
    from types import SimpleNamespace

    def make_args(**overrides):
        kwargs = dict(DEFAULT_SCREEN_ARGS)
        kwargs["period"] = period
        kwargs.update(overrides)
        return SimpleNamespace(**kwargs)

    jobs = []
    if theme_kw:
        for kw in [k.strip() for k in theme_kw.split(",") if k.strip()]:
            jobs.append((f"테마_{kw}"[:31], make_args(theme_keyword=kw)))

    if wl_input:
        jobs.append(("관심종목", make_args(code=wl_input, name=wl_input, buy_price=buy_price, quantity=quantity)))
    elif not theme_kw:
        # 통합조회의 중심은 테마주이므로, 관심종목은 테마 없이 단독 실행될 때만
        # 직접 입력이 없어도 관심종목.txt를 기본으로 조회한다.
        jobs.append(("관심종목", make_args(watchlist=watchlist_path, buy_price=buy_price, quantity=quantity)))
    if search_kw:
        jobs.append(("종목검색", make_args(search=search_kw, buy_price=buy_price, quantity=quantity)))
    if lookup:
        jobs.append(("종목조회", make_args(code=lookup, name=lookup, buy_price=buy_price, quantity=quantity)))

    sheets = []
    for sheet_name, args in jobs:
        print(f"\n=== [{sheet_name}] 조회 중 ===", file=sys.stderr)
        try:
            sheets.append((sheet_name, run_screen(args)))
        except (FileNotFoundError, ValueError) as e:
            print(f"[{sheet_name}] 건너뜀: {e}", file=sys.stderr)
    return sheets


def main():
    ap = argparse.ArgumentParser(description="한국 주식 조건검색 스크리너 (네이버금융 기반, 로그인 불필요)")
    ap.add_argument("--period", type=int, default=20, help="누적하락/상승일 계산 기간(거래일 수, 기본 20)")
    ap.add_argument("--market", default="ALL", choices=["ALL", "KOSPI", "KOSDAQ"], help="시장 구분")

    ap.add_argument("--market-cap-min", type=float, default=0, help="시가총액 하한 (억원)")
    ap.add_argument("--volume-min", type=float, default=0, help="당일 거래량 하한 (주)")
    ap.add_argument("--cum-decline-min", type=float, default=None, help="누적하락률 하한 (%%, 예: 20 -> 기간중 20%% 이상 하락)")
    ap.add_argument("--up-days-min", type=int, default=None, help="기간 내 상승일수 하한")
    ap.add_argument("--consecutive-up-min", type=int, default=None, help="최근 연속 상승일수 하한")
    ap.add_argument("--theme-keyword", default=None, help="이슈테마주 키워드 (콤마로 복수 지정, 네이버금융 테마 기준)")

    ap.add_argument("--min-data-days", type=int, default=None,
                     help="분석하한: 최소 거래 이력 일수. 미충족 종목(신규상장 등)은 제외. 기본값=period")
    ap.add_argument("--code", default=None,
                     help="특정 종목만 조회 (종목코드, 콤마로 복수 지정). 예: 006400 (삼성SDI). "
                          "지정 시 시총/거래량 하한 필터는 건너뛰고 해당 종목의 지표만 계산합니다.")
    ap.add_argument("--name", default=None,
                     help="특정 종목만 조회 (종목명, 콤마로 복수 지정). 예: 삼성SDI. --code와 동일하게 동작합니다.")
    ap.add_argument("--search", default=None,
                     help="종목명에 이 글자가 포함된 종목을 모두 검색 (정확한 이름을 몰라도 됨). "
                          "예: --search 삼성 -> 삼성전자/삼성SDI/삼성물산 등 전부 검색.")
    ap.add_argument("--watchlist", default=None,
                     help="관심종목 텍스트 파일 경로. 파일에 적힌 종목들을 --code/--name과 동일하게 조회합니다. "
                          "예: --watchlist 관심종목.txt")
    ap.add_argument("--out", default="screener_result.xlsx",
                     help="결과 엑셀 파일명 (.xlsx로 저장됨, 확장자는 자동으로 맞춰짐)")
    ap.add_argument("--no-open", action="store_true",
                     help="결과 엑셀 파일을 자동으로 열지 않음 (기본값: 완료 후 엑셀로 자동으로 엶)")
    ap.add_argument("--buy-price", type=float, default=None,
                     help="평균단가(매입 평균단가). 지정하면 보유잔고 화면 형태(평가손익/손익률/매입금/평가금/"
                          "매도가능/평균단가/보유잔고/현재가)로 표가 바뀌고, 엑셀에 손익 그래프도 삽입됩니다.")
    ap.add_argument("--quantity", type=int, default=None,
                     help="보유잔고(수량). --buy-price와 함께 지정. 생략하면 1주로 계산합니다.")
    ap.add_argument("--combined", action="store_true",
                     help="관심종목/종목검색/종목조회/테마주를 한 번에 조회해서 "
                          "한 엑셀 파일의 각 시트로 저장하고, 끝나면 그 파일 하나만 엽니다.")
    ap.add_argument("--wl-input", default=None,
                     help="(--combined 전용) 관심종목 직접 입력 (콤마로 여러개). 생략하면 --watchlist 파일 사용.")
    ap.add_argument("--search-kw", default=None, help="(--combined 전용) 종목검색 키워드")
    ap.add_argument("--lookup", default=None, help="(--combined 전용) 종목조회 종목코드/종목명")

    args = ap.parse_args()

    if args.combined:
        try:
            sheets = run_combined(
                wl_input=args.wl_input,
                watchlist_path=args.watchlist or "관심종목.txt",
                search_kw=args.search_kw,
                lookup=args.lookup,
                theme_kw=args.theme_keyword,
                buy_price=args.buy_price,
                quantity=args.quantity,
                period=args.period,
            )
        except ValueError as e:
            print(str(e), file=sys.stderr)
            sys.exit(1)

        if not sheets:
            print("모든 항목에서 결과가 없습니다.", file=sys.stderr)
            sys.exit(1)

        out_path = os.path.splitext(args.out)[0] + ".xlsx"
        saved_path = save_excel_sheets(sheets, out_path)
        print(f"\n완료: {len(sheets)}개 시트, 결과 저장 -> {saved_path}", file=sys.stderr)
        for name, df in sheets:
            print(f"\n[{name}] {len(df)}종목")
            print(format_table(df.head(15)))

        if not args.no_open:
            try:
                os.startfile(saved_path)
            except Exception:
                pass
        return

    try:
        result = run_screen(args)
    except (FileNotFoundError, ValueError) as e:
        print(str(e), file=sys.stderr)
        sys.exit(1)

    out_path = os.path.splitext(args.out)[0] + ".xlsx"
    saved_path = save_excel(result, out_path)
    print(f"\n완료: {len(result)}종목, 결과 저장 -> {saved_path}", file=sys.stderr)
    print(format_table(result.head(30)))

    if not args.no_open:
        try:
            os.startfile(saved_path)
        except Exception:
            pass


if __name__ == "__main__":
    main()
